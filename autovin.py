import streamlit as st
import pandas as pd
import requests
import os
import openpyxl
import numpy as np
import json
from datetime import datetime
from io import BytesIO
import warnings

@st.cache
def confirm_vin(file):
    # Handle Excel files with multiple sheets
    wb = openpyxl.load_workbook(file)
    if len(wb.sheetnames) > 1:
        raw_vin_data = pd.read_excel(file, 'Vehicle & Asset List', header=3)
    else:
        raw_vin_data = pd.read_excel(file, header=3)

    # Standardize column names
    for column in raw_vin_data.columns:
        if 'vehicle asset name' in column.lower():
            raw_vin_data.rename(columns={column: 'Vehicle Asset Name'}, inplace=True)
        elif 'model year' in column.lower():
            raw_vin_data.rename(columns={column: 'Model Year'}, inplace=True)
        elif 'make' in column.lower():
            raw_vin_data.rename(columns={column: 'Make'}, inplace=True)
        elif 'model' in column.lower():
            raw_vin_data.rename(columns={column: 'Model'}, inplace=True)
        elif 'vin' in column.lower():
            raw_vin_data.rename(columns={column: 'VIN'}, inplace=True)
        elif 'fuel type' in column.lower():
            raw_vin_data.rename(columns={column: 'Fuel Type'}, inplace=True)

    base_url = 'https://vpic.nhtsa.dot.gov/api/vehicles/DecodeVin/'
    vin_data = pd.DataFrame({'VRN': [], 'VIN': [], 'YEAR': [], 'MAKE': [], 'MODEL': [], 'FUEL': [], 'COUNTRY': []})

    for ind in raw_vin_data.index:
        if pd.notna(raw_vin_data['VIN'][ind]):
            vin_data.loc[ind] = [raw_vin_data['Vehicle Asset Name'][ind],
                                 raw_vin_data['VIN'][ind], raw_vin_data['Model Year'][ind],
                                 raw_vin_data['Make'][ind], raw_vin_data['Model'][ind],
                                 raw_vin_data['Fuel Type'][ind], 'US']

    vin_data.reset_index(drop=True, inplace=True)
    vin_data.replace(np.nan, '', inplace=True)
    vin_data = vin_data.astype(str)

    results = []
    values = vin_data['VIN'].values.tolist()
    ind = 0

    for value in values:
        value = str(value).replace(" ", "")
        url = base_url + value + '?format=json'
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", requests.packages.urllib3.exceptions.InsecureRequestWarning)
            response = requests.get(url, verify=False)
        
        try:
            data = response.json()
            decoded_values = {item['Variable']: item['Value'] for item in data['Results']}
            results.append({
                'VRN': vin_data['VRN'][ind],
                'VIN': value,
                'YEAR': decoded_values.get('Model Year', 'N/A'),
                'MAKE': decoded_values.get('Make', 'N/A'),
                'MODEL': decoded_values.get('Model', 'N/A'),
                'FUEL': decoded_values.get('Fuel Type - Primary', 'N/A'),
                'COUNTRY': 'US',
                'VEHICLE TYPE': decoded_values.get('Vehicle Type', 'N/A'),
                'ERROR CODE': decoded_values.get('Error Text', 'N/A')
            })
            ind += 1
        except json.JSONDecodeError:
            results.append({
                'VRN': vin_data['VRN'][ind],
                'VIN': value,
                'YEAR': 'Error',
                'MAKE': 'Error',
                'MODEL': 'Error',
                'FUEL': 'Error',
                'COUNTRY': 'Error',
                'VEHICLE TYPE': 'Error',
                'ERROR CODE': 'Error: No information found for input VIN'
            })
            ind += 1
        except requests.exceptions.Timeout:
            return "Timed out"

    results = pd.DataFrame(results)
    valid_vins = results[~results.FUEL.isin(['Not Applicable', 'Error', None])]
    valid_vins = valid_vins[datetime.now().year - valid_vins['YEAR'].astype(int) < 30]
    valid_vins.drop(['VEHICLE TYPE', 'ERROR CODE'], axis=1, inplace=True)
    valid_vins.drop_duplicates(subset=['VIN'], inplace=True)

    check_list = []
    vins_checked = []
    valid_vin_list = valid_vins['VIN'].values.tolist()
    vin_data = pd.concat([vin_data, results['VEHICLE TYPE']], axis=1)

    for ind in vin_data.index:
        if vin_data['VIN'][ind].replace(" ", "") in valid_vin_list and vin_data['VIN'][ind] not in vins_checked:
            check_list.append('NO')
        elif vin_data['VEHICLE TYPE'][ind] == 'TRAILER':
            check_list.append('NO')
        elif 'trailer' in vin_data['MODEL'][ind].lower() or 'trailer' in vin_data['VRN'][ind].lower():
            check_list.append('NO')
        elif 'lift' in vin_data['MODEL'][ind].lower() or 'lift' in vin_data['VRN'][ind].lower():
            check_list.append('NO')
        elif 'example' in vin_data['VIN'][ind].lower():
            check_list.append('NO')
        elif vin_data['VIN'][ind] in vins_checked:
            check_list.append('YES: Duplicate Vin')
        else:
            check_list.append('YES')
        vins_checked.append(vin_data['VIN'][ind])

    for ind in vin_data.index:
        if vin_data['VEHICLE TYPE'][ind] == None or vin_data['VEHICLE TYPE'][ind] == 'Error':
            if 'trailer' in vin_data['MODEL'][ind].lower():
                vin_data['VEHICLE TYPE'][ind] = 'TRAILER'
            elif 'lift' in vin_data['MODEL'][ind].lower():
                vin_data['VEHICLE TYPE'][ind] = 'LIFT'
            elif vin_data['VEHICLE TYPE'][ind] == 'Error':
                vin_data['VEHICLE TYPE'][ind] = 'UNKNOWN'

    vin_data['MANUAL CHECK NEEDED'] = check_list
    vin_data = pd.concat([vin_data, results['ERROR CODE']], axis=1)

    CAN_file_path = os.path.splitext(file.name)[0] + "_CAN.csv"
    pd.DataFrame(valid_vins).to_csv(CAN_file_path, index=False)

    processed_file_path = os.path.splitext(file.name)[0] + "_processed.xlsx"
    with pd.ExcelWriter(processed_file_path, engine='openpyxl') as writer:
        vin_data.to_excel(writer, index=False, sheet_name='Processed VINs')
        workbook = writer.book
        worksheet = writer.sheets['Processed VINs']

        for idx, column in enumerate(worksheet.columns):
            if worksheet.cell(row=1, column=idx + 1).value != 'ERROR CODE':
                max_length = 0
                for cell in column:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[chr(65 + idx)].width = adjusted_width
            
            if worksheet.cell(row=1, column=idx + 1).value == 'ERROR CODE':
                worksheet.column_dimensions[chr(65 + idx)].width = 12

    return processed_file_path, CAN_file_path

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css?family=Open+Sans');
body {
    font-family: 'Open Sans', sans-serif;
}
</style>
""", unsafe_allow_html=True)

st.image("https://www.tdtyres.com/wp-content/uploads/2018/12/kisspng-car-michelin-man-tire-logo-michelin-logo-5b4c286206fa03.5353854915317177300286.png")
st.title("VIN Decoder")

uploaded_file = st.file_uploader("Choose an Excel or CSV file", type=["xls", "xlsx", "csv"])

if "processed_file_path" not in st.session_state:
    st.session_state["processed_file_path"] = None
    st.session_state["can_file_path"] = None

if uploaded_file is not None:
    with st.spinner('Processing...'):
        processed_file_path, can_file_path = confirm_vin(uploaded_file)
        st.session_state["processed_file_path"] = processed_file_path
        st.session_state["can_file_path"] = can_file_path
        st.success('File successfully processed!')

if st.session_state["processed_file_path"] and st.session_state["can_file_path"]:
    with open(st.session_state["processed_file_path"], "rb") as f:
        processed_data = f.read()
    with open(st.session_state["can_file_path"], "rb") as f:
        can_data = f.read()
    st.download_button(
        label="Download Processed File",
        data=BytesIO(processed_data),
        file_name=os.path.basename(st.session_state["processed_file_path"]),
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    st.download_button(
        label="Download CAN File",
        data=BytesIO(can_data),
        file_name=os.path.basename(st.session_state["can_file_path"]),
        mime='text/csv'
    )

st.markdown('''
This application checks customer VINs with the [National Highway Traffic Safety Administration API](https://vpic.nhtsa.dot.gov/api/) to confirm VIN accuracy. The API helps ensure the VINs are accurate and relate to relevant vehicles for the CAN compatibility check on Salesforce. This application can handle large volumes of VINs but greater numbers of uploaded VINs will slow down processing time. Processing 2200 VINs takes roughly 25 minutes. When uploading large numbers of VINs please be patient and do not close out the application while processing.

**Input Document Requirements**

- The uploaded document containing the VINs must follow the standard [Michelin Connected Fleet Deployment Template.](https://michelingroup.sharepoint.com/:x:/r/sites/ProcessImprovement/_layouts/15/Doc.aspx?sourcedoc=%7BFA264B31-B424-418C-8D1C-C0E5F001094E%7D&file=MCF%20Deployment%20Template.xlsx&action=default&mobileredirect=true&wdsle=0) This application cannot decipher different document formats. If an error is indicated with a file you upload, please check the uploaded document follows the formatting guidelines.
- The VIN column must include the VINs the user wants to query. This is the only field necessary to confirm the existence/accuracy of the VINs.
- The output documents will lack information regarding the vehicle make, model, year, and fuel type if these input document columns are empty. If you are interested in retrieving vehicle make, model, year, etc. information from VINs alone please use the [Automated VIN Data Application](https://vindata.streamlit.app/).

***Example Input Document:*** [***VIN Example***](https://michelingroup.sharepoint.com/:x:/r/sites/ProcessImprovement/_layouts/15/Doc.aspx?sourcedoc=%7B58E5DF8A-9843-481F-A3E6-16A6B422D4EC%7D&file=VIN%20Example.xlsx&action=default&mobileredirect=true&wdsle=0)

***Note:*** If you are interested in checking the accuracy/existence of VINs recorded in a different format/document download the MCF Deployment Template linked above, then copy and paste the VINs into the VIN column and upload this document for bulk processing.
''')

st.markdown('''
**Output File 1: CAN Compatibility Check**
- After comparison with the NHTSA VIN database, accurate and relevant VINs are written to a CSV file following the standard format for the CAN compatibility check.
- VINs relating to trailers and lifts are considered irrelevant to the CAN compatibility check and are excluded from this document.
- This CSV will have the same name as the original document followed by _CAN. This file includes VRN, Year, Make, Model, VIN, and Fuel Type information from the original input file.

***Example CAN Output Document:*** [***VIN Example_CAN***](https://michelingroup.sharepoint.com/:x:/r/sites/ProcessImprovement/_layouts/15/Doc.aspx?sourcedoc=%7BF10B0EAE-4CAF-4F19-9B61-BB1F62DEDFC0%7D&file=VIN%20Example_CAN.xlsx&action=default&mobileredirect=true&wdsle=0)
''')

st.markdown('''
**Output File 2: Processed VINs**
- This secondary output file includes information on all VINs present in the original uploaded document, including VINs excluded from the CAN Compatibility Check document.
- The application processes the original VIN document and determines the VIN's vehicle type, indicates whether a manual employee check for a VIN is necessary and provides error code information pertaining to the VIN.
- An error code of 0 indicates there was no issue with the VIN.
- A manual check is indicated as unnecessary if the VIN was considered valid and written to the CAN compatibility document or the vehicle type is a trailer or lift (irrelevant vehicles).
- A manual check is necessary if the VIN was not written to the CAN compatibility file as a valid VIN and the VIN does not relate to a trailer or lift (could be a relevant vehicle).
- This file will have the same name as the original document followed by _processed. This file also includes VRN, Year, Make, Model, VIN and Fuel Type information from the original document.

***Example Processed Output Document:*** [***VIN Example_processed***](https://michelingroup.sharepoint.com/:x:/r/sites/ProcessImprovement/_layouts/15/Doc.aspx?sourcedoc=%7B56DE5CED-7E83-459B-9430-BF55C85CD22A%7D&file=VIN%20Example_processed.xlsx&action=default&mobileredirect=true&wdsle=0)

If you are encountering issues with this application please contact Massaer Diouf.
''')
